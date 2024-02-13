import boto3
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import json

# Initialize AWS clients
org_client = boto3.client('organizations')
securityhub_client = boto3.client('securityhub')
s3_client = boto3.client('s3')

def list_all_accounts():
    """
    Fetches all AWS accounts in the organization.
    :return: List of account IDs.
    """
    account_ids = []
    paginator = org_client.get_paginator('list_accounts')
    for page in paginator.paginate():
        for account in page['Accounts']:
            account_ids.append(account['Id'])
    return account_ids

def get_account_owner_email_and_name(account_id):
    """
    Retrieves the email of the account owner and the account name.
    :param account_id: AWS Account ID to fetch details for.
    :return: Tuple containing the email and account name.
    """
    response = org_client.describe_account(AccountId=account_id)
    account_name = response['Account']['Name']
    tags = org_client.list_tags_for_resource(ResourceId=account_id)
    email = None
    for tag in tags['Tags']:
        if tag['Key'] == 'Account-Owner':
            email = tag['Value']
            break
    return email, account_name

def get_account_findings(account_id, filters):
    """
    Retrieves findings for a specific AWS account.
    :param account_id: AWS Account ID.
    :param filters: Filters to apply to the Security Hub findings query.
    :return: List of findings for the account.
    """
    
    ########### update filters to include the account id ######
    filters['AwsAccountId'] = [{'Value': account_id, 'Comparison': 'EQUALS'}]
    ###########################################################

    findings = []
    paginator = securityhub_client.get_paginator('get_findings')
    for page in paginator.paginate(Filters=filters):
        findings.extend([f for f in page['Findings'] if f['AwsAccountId'] == account_id])
    
    return findings

def create_consolidated_excel_report(all_findings):
    """
    Creates a consolidated Excel report for all accounts.
    :param all_findings: List of all findings across accounts.
    :return: BytesIO object containing the Excel report.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'All Accounts Findings'

    # Headers
    headers = ['Account ID', 'Account Name', 'Account Owner Email', 'Finding Title', 'Severity', 'Resource ID', 'Region']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Populate the sheet with findings
    for row_num, finding in enumerate(all_findings, 2):
        sheet.cell(row=row_num, column=1).value = finding['account_id']
        sheet.cell(row=row_num, column=2).value = finding['account_name']
        sheet.cell(row=row_num, column=3).value = finding['owner_email']
        sheet.cell(row=row_num, column=4).value = finding.get('Title')
        sheet.cell(row=row_num, column=5).value = finding.get('Severity', {}).get('Label')
        sheet.cell(row=row_num, column=6).value = finding.get('Resources', [{}])[0].get('Id')
        sheet.cell(row=row_num, column=7).value = finding.get('Region')

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except TypeError:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save to BytesIO object
    excel_data = io.BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)
    return excel_data.getvalue()

def lambda_handler(event, context):
    date_str = datetime.now().strftime('%Y-%m-%d')
    folder_name = f'{date_str}/'

    accounts = list_all_accounts()

    owner_account_info = {}  # Group accounts by owner

    for account_id in accounts:
        email, account_name = get_account_owner_email_and_name(account_id)
        if email:
            if email not in owner_account_info:
                owner_account_info[email] = []
            owner_account_info[email].append({'account_id': account_id, 'account_name': account_name})
            
    all_findings = []  # List to hold findings from all accounts

    # Process each owner's accounts
    for owner_email, accounts in owner_account_info.items():
        for account in accounts:
            account_id = account['account_id']
            # Define filters
            filters = {
                'ComplianceStatus': [{'Value': 'FAILED', 'Comparison': 'EQUALS'}],
                'SeverityLabel': [{'Value': 'CRITICAL', 'Comparison': 'EQUALS'}, {'Value': 'HIGH', 'Comparison': 'EQUALS'}],
                'WorkflowStatus': [{'Value': 'NEW', 'Comparison': 'EQUALS'}, {'Value': 'NOTIFIED', 'Comparison': 'EQUALS'}],
                'RecordState': [{'Value': 'ACTIVE', 'Comparison': 'EQUALS'}]
            }

            findings = get_account_findings(account_id, filters)

            for finding in findings:
                finding_data = {
                    'account_id': account_id,
                    'account_name': account['account_name'],
                    'owner_email': owner_email,
                    'Title': finding.get('Title'),
                    'Severity': finding.get('Severity'),
                    'Resources': finding.get('Resources'),
                    'Region': finding.get('Region')
                }
                all_findings.append(finding_data)

    excel_content = create_consolidated_excel_report(all_findings)
    file_name = f"All_Findings_{date_str}.xlsx"
    s3_client.put_object(Bucket='medina-awsplatform-account-owner-reports', Key=folder_name + file_name, Body=excel_content)

    return {
        'statusCode': 200,
        'body': json.dumps('Consolidated report generated successfully!')
    }

